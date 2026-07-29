"""
Laghaim Original Rankings & Daily XP Scraper (Web Dashboard Ready)
Scrapes rankings and daily XP, maintains 7-day rolling XP history,
calculates snapshot deltas, and outputs both data.json and LaghaimRankings.xlsx.
"""

from datetime import datetime, timedelta
import json
from pathlib import Path
import sys
import time
from typing import Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd
from playwright.sync_api import TimeoutError, sync_playwright

# ==========================================================
# SETTINGS & CONFIGURATION
# ==========================================================

URL = "https://www.laghaim-original.com/ranking.xhtml"
REPORT_URL = "https://www.laghaim-original.com/report_ranking.xhtml"

HEADLESS = False
RETRY_COUNT = 3

OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

EXCEL_FILE = OUTPUT_DIR / "LaghaimRankings.xlsx"
JSON_FILE = OUTPUT_DIR / "data.json"


# ==========================================================
# FILE LOCK SAFEGUARDS
# ==========================================================

def check_file_lock(filepath: Path) -> None:
    """Pre-flight check: Verifies if the target Excel file is locked."""
    if not filepath.exists():
        return

    while True:
        try:
            with open(filepath, "r+"):
                break
        except IOError:
            print(f"\n[WARNING] Cannot access '{filepath.name}' — open in Excel?")
            response = input("Close the file and press ENTER to continue (or 'q' to quit): ").strip().lower()
            if response == 'q':
                sys.exit(0)


def safe_save_writer(writer: pd.ExcelWriter, filepath: Path) -> None:
    """Post-scrape safeguard for Pandas ExcelWriter."""
    while True:
        try:
            writer.close()
            break
        except PermissionError:
            print(f"\n[ERROR] Permission denied: Unable to save to '{filepath.name}'.")
            response = input("Please close Excel and press ENTER to try saving again: ").strip().lower()
            if response == 'q':
                sys.exit(1)


def safe_save_workbook(wb, filepath: Path) -> None:
    """Post-formatting safeguard for OpenPyXL workbook."""
    while True:
        try:
            wb.save(filepath)
            break
        except PermissionError:
            print(f"\n[ERROR] Unable to save workbook formatting to '{filepath.name}'.")
            response = input("Please close Excel and press ENTER to try saving again: ").strip().lower()
            if response == 'q':
                sys.exit(1)


# ==========================================================
# SCRAPE FUNCTIONS
# ==========================================================

def scrape_page(page) -> List[Dict]:
    """Scrapes all player records from the current rankings table page."""
    players = []
    rows = page.locator("table tbody tr")

    for r in range(rows.count()):
        row = rows.nth(r)
        cols = row.locator("td")

        if cols.count() < 8:
            continue

        raw_exp = cols.nth(5).inner_text().strip().replace(".", "")

        players.append({
            "Rank": int(cols.nth(0).inner_text().strip()),
            "Name": cols.nth(1).inner_text().strip(),
            "Level": int(cols.nth(3).inner_text().strip()),
            "Guild": cols.nth(4).inner_text().strip(),
            "Experience": int(raw_exp),
            "Race": cols.nth(6).inner_text().strip(),
        })

    return players


def wait_for_report_table(page) -> None:
    """Waits until daily report table count stabilizes."""
    previous_count = 0
    stable_count = 0

    while stable_count < 5:
        count = page.locator("table tbody tr").count()
        if count == previous_count and count > 0:
            stable_count += 1
        else:
            stable_count = 0
        previous_count = count
        time.sleep(1)


def scrape_daily_xp() -> pd.DataFrame:
    """Navigates to report page, sets yesterday's date, and collects daily XP."""
    players = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        page = browser.new_page(viewport={"width": 1600, "height": 900})

        print("Opening XP report...")
        page.goto(REPORT_URL)
        page.wait_for_load_state("networkidle")

        server_date = datetime.now(ZoneInfo("Europe/Berlin"))
        report_date = (server_date - timedelta(days=1)).strftime("%Y-%m-%d")

        print(f"Changing XP report date to: {report_date}")
        page.locator("#report_date").fill(report_date)
        page.locator("#report_date").press("Enter")

        wait_for_report_table(page)

        rows = page.locator("table tbody tr")
        print(f"XP rows found: {rows.count()}")

        for r in range(rows.count()):
            cols = rows.nth(r).locator("td")
            if cols.count() < 4:
                continue

            rank_text = cols.nth(0).inner_text().strip()
            if not rank_text.isdigit():
                continue

            raw_xp = cols.nth(3).inner_text().strip().replace(".", "")

            players.append({
                "Date": report_date,
                "Name": cols.nth(1).inner_text().strip(),
                "Daily XP": int(raw_xp) if raw_xp.isdigit() else 0,
            })

        browser.close()

    return pd.DataFrame(players)


def load_previous_history() -> Optional[pd.DataFrame]:
    """Loads previous snapshot from Rankings_Old sheet."""
    if not EXCEL_FILE.exists():
        return None
    try:
        return pd.read_excel(EXCEL_FILE, sheet_name="Rankings_Old")
    except Exception:
        return None


def format_change(value) -> str:
    """Formats numeric differences into directional indicator strings."""
    if pd.isna(value):
        return "★ NEW"
    elif value > 0:
        return f"▲ +{int(value)}"
    elif value < 0:
        return f"▼ {int(value)}"
    else:
        return "="


def create_changes(
    current: pd.DataFrame, previous: Optional[pd.DataFrame]
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Compares current rankings against previous snapshots."""
    if previous is None or previous.empty:
        empty = pd.DataFrame()
        return empty, empty, empty, empty

    comparison = current.merge(
        previous,
        on="Name",
        how="outer",
        suffixes=("", "_Previous"),
        indicator=True,
    )

    new_players = comparison[comparison["_merge"] == "left_only"][["Name", "Rank"]]
    removed_players = comparison[comparison["_merge"] == "right_only"][["Name", "Rank_Previous"]]

    changes = comparison[comparison["_merge"].isin(["both", "left_only"])].copy()

    changes["Rank Change"] = changes["Rank_Previous"] - changes["Rank"]
    changes["Level Change"] = changes["Level"] - changes["Level_Previous"]

    changes["Rank Movement"] = changes["Rank Change"].apply(format_change)
    changes["Level Movement"] = changes["Level Change"].apply(format_change)

    changes["Rank_Previous"] = changes["Rank_Previous"].fillna("NEW")
    changes["Level_Previous"] = changes["Level_Previous"].fillna("-")
    changes["Guild_Previous"] = changes["Guild_Previous"].fillna("")

    guild_changes = comparison[
        (comparison["_merge"] == "both") & (comparison["Guild"] != comparison["Guild_Previous"])
    ][["Name", "Guild_Previous", "Guild"]]

    changes = changes[[
        "Name",
        "Rank",
        "Rank_Previous",
        "Rank Movement",
        "Level",
        "Level_Previous",
        "Level Movement",
        "Guild",
        "Guild_Previous",
    ]].sort_values("Rank")

    return changes, new_players, removed_players, guild_changes


# ==========================================================
# MAIN EXECUTION PIPELINE
# ==========================================================

if __name__ == "__main__":

    check_file_lock(EXCEL_FILE)

    all_players = []

    # 1. Scrape Main Rankings Table
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        page = browser.new_page(viewport={"width": 1600, "height": 900})

        print("Opening rankings website...")
        page.goto(URL)
        page.wait_for_load_state("networkidle")

        current_page = 1

        while True:
            retries = RETRY_COUNT
            while retries > 0:
                try:
                    rows = scrape_page(page)
                    print(f"Page {current_page:>3} : {len(rows)} players")
                    all_players.extend(rows)
                    break
                except Exception as e:
                    retries -= 1
                    print(f"Retrying page {current_page}: {e}")

            next_btn = page.locator("a.ui-paginator-next")
            btn_class = next_btn.get_attribute("class") or ""

            if "ui-state-disabled" in btn_class:
                print("Reached last ranking page.")
                break

            first_td = page.locator("table tbody tr td:nth-child(2)").first
            old_name = first_td.inner_text() if first_td.count() > 0 else ""

            next_btn.click(force=True)

            try:
                page.wait_for_function(
                    """
                    oldName => {
                        const td = document.querySelector('table tbody tr td:nth-child(2)');
                        return td && td.innerText.trim() !== oldName;
                    }
                    """,
                    arg=old_name,
                    timeout=10000,
                )
            except TimeoutError:
                print("Timed out waiting for ranking page transition.")
                break

            current_page += 1

        browser.close()

    # 2. Scrape Daily XP Report
    daily_xp = scrape_daily_xp()

    # 3. Process Main Rankings DataFrame
    df = pd.DataFrame(all_players)
    df.drop_duplicates(subset=["Name"], inplace=True)
    df.sort_values("Rank", inplace=True)

    # 4. Prepare Daily XP & 7-Day Rolling History
    daily_xp["Daily XP"] = daily_xp["Daily XP"].fillna(0).astype(int)
    xp_history_today = daily_xp[["Date", "Name", "Daily XP"]].copy()

    if EXCEL_FILE.exists():
        try:
            old_xp_history = pd.read_excel(EXCEL_FILE, sheet_name="XP History")
        except Exception:
            old_xp_history = pd.DataFrame()
    else:
        old_xp_history = pd.DataFrame()

    xp_history = pd.concat([old_xp_history, xp_history_today], ignore_index=True)
    xp_history = xp_history.drop_duplicates(subset=["Date", "Name"], keep="last")

    xp_history["Date"] = pd.to_datetime(xp_history["Date"])

    # Enforce 7-Day Rolling Window
    max_date = xp_history["Date"].max()
    xp_history = xp_history[xp_history["Date"] >= (max_date - timedelta(days=6))]

    # Calculate 7-Day Average (excluding active 0 XP days)
    xp_average = (
        xp_history[xp_history["Daily XP"] > 0]
        .groupby("Name")["Daily XP"]
        .mean()
        .round(0)
        .astype(int)
        .reset_index(name="Average Daily XP")
    )

    xp_history["Date"] = xp_history["Date"].dt.strftime("%Y-%m-%d")

    # 5. Merge XP Data into Main Dataset
    df = df.merge(daily_xp[["Name", "Daily XP"]], on="Name", how="left")
    df["Daily XP"] = df["Daily XP"].fillna(0).astype(int)

    df = df.merge(xp_average, on="Name", how="left")
    df["Average Daily XP"] = df["Average Daily XP"].fillna(0).astype(int)

    # 6. Generate Deltas
    previous_rankings = load_previous_history()
    changes, new_players, removed_players, guild_changes = create_changes(
        df, previous_rankings
    )

    # Preserve Prior Baseline
    if EXCEL_FILE.exists():
        try:
            old_rankings = pd.read_excel(EXCEL_FILE, sheet_name="Rankings")
        except Exception:
            old_rankings = pd.DataFrame()
    else:
        old_rankings = pd.DataFrame()

    # ==========================================================
    # EXPORT DATA.JSON (For Web Dashboard)
    # ==========================================================
    web_export = {
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "rankings": df.to_dict(orient="records"),
        "changes": changes.to_dict(orient="records") if not changes.empty else [],
        "xp_history": xp_history.to_dict(orient="records"),
    }

    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(web_export, f, indent=2)

    print(f"\n[OK] Output exported to JSON: {JSON_FILE}")

    # ==========================================================
    # EXPORT RAW EXCEL FILE (For Local Backup/Reference)
    # ==========================================================
    writer = pd.ExcelWriter(EXCEL_FILE, engine="openpyxl")

    df.to_excel(writer, sheet_name="Rankings", index=False)
    xp_history.to_excel(writer, sheet_name="XP History", index=False)
    changes.to_excel(writer, sheet_name="Changes", index=False)

    if not old_rankings.empty:
        old_rankings.to_excel(writer, sheet_name="Rankings_Old", index=False)

    safe_save_writer(writer, EXCEL_FILE)

    # Clean OpenPyXL auto-formatting for the 4 backup sheets
    wb = load_workbook(EXCEL_FILE)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.auto_filter.ref = ws.dimensions
        for column_cells in ws.columns:
            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            col_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[col_letter].width = max(length + 3, 10)

    safe_save_workbook(wb, EXCEL_FILE)

    print("==========================")
    print("Scrape & Export Completed!")
    print("==========================")